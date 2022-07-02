from openpyxl import load_workbook
from os import listdir
from pprint import pprint

# NOTE: The arrows(->) show the return type of a non void function. This doesn't affect the program and is intended for documentation purposes only.
# NOTE: The timtables provided in the project are test timetables only. It is possible to interface this with any timetable. 🙂


def readWorkbook(path) -> dict:
    wb = load_workbook(path)
    ws = wb.active
    num = ws.max_row
    compiled = {}
    for i in range(2, num + 1):
        lis = [i.value for i in ws[i]]
        compiled[lis[0]] = lis[1:]
    return compiled


# pprint(readWorkbook(
#     r"C:\Users\abhin\OneDrive\Desktop\SchoolData\Bhaswati Chattopadhyay.xlsx")
# )

# 2 kinds of clashes:
# 1. Both teachers assigned to same period
# 2. One teacher assigned to 2 places at once


def checkClashes(path1, path2, day) -> list:
    # Checks clashes if both teachers are assigned to one period, accounting for the fact that they teach different streams
    t1 = readWorkbook(path1)
    t2 = readWorkbook(path2)
    day1 = t1[day]
    day2 = t2[day]
    clashes = []
    for i in range(len(day1)):
        if day1[i] and day2[i]:
            p1 = day1[i].split('\n')
            p2 = day2[i].split('\n')
            for j in p1:
                if j in p2:
                    clashes.append(
                        f'Period {i + 1}: Clash detected. 2 teachers assigned in the same period for 1 or more groups.')
                    break

    return clashes


# res = checkClashes(r"SchoolData\Bhaswati Chattopadhyay.xlsx",
#                    r"SchoolData\Bini P Kuriakose.xlsx", 'Monday')

# pprint(res)


def viewFreeAndBusy(folder, day, period_number, view_busy=False) -> list:
    files = listdir(folder)
    freeTeachers = []
    busyTeachers = []
    for i in files:
        period = readWorkbook(f'{folder}\{i}')[
            day][period_number - 1]
        if period == None:
            freeTeachers.append(i)
        else:
            busyTeachers.append(i)

    if view_busy == False:
        return freeTeachers
    return busyTeachers


def style_worksheet(ws, cell_range):  # Function used to style the worksheet
    from openpyxl.styles import Border, Side, Alignment
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            cell.alignment = Alignment(horizontal='center')


# Unused args: teacher, subject, number_of_days, periods_in_a_day
def createNewTimetable(number_of_periods, number_of_days, folder, name, teacher_and_period, allPeriods=False):
    import string
    import xlsxwriter
    from openpyxl import load_workbook
    import random
    from pprint import pprint
    from copy import deepcopy

    colLetters = [letter for letter in string.ascii_uppercase]

    days_of_the_week = ['Monday', 'Tuesday', 'Wednesday',
                        'Thursday', 'Friday', 'Saturday', 'Sunday']

    worksheet_col1 = ['Days'] + days_of_the_week[:number_of_days]

    workbook = xlsxwriter.Workbook(rf'{folder}/{name}.xlsx')
    worksheet = workbook.add_worksheet('Timetable')

    for i in range(number_of_days + 1):
        worksheet.write(i, 0, worksheet_col1[i])

    for i in range(1, number_of_periods + 1):
        # Essentially this for loop is writing the period number in the first row of the timetable sheet.
        worksheet.write(0, i, i)
    workbook.close()

    # All the days in which school is running
    schoolDays = days_of_the_week[:number_of_days]

    wb = load_workbook(filename=rf'{folder}/{name}.xlsx')
    ws = wb.active

    subjects = list(teacher_and_period.keys())

    raw_table = {}

    for day in schoolDays:
        i = 0
        periods = []
        available_periods = deepcopy(subjects)
        while i < number_of_periods:
            period = random.choice(available_periods)
            numPeriod = periods.count(period)
            if numPeriod <= 1:  # This works but i don't know why. For some reason, this creates a timetable in which a period can occur at max 2 times a day.
                periods.append(period)
            else:
                available_periods.remove(period)
                continue
            i += 1
        raw_table[day] = periods

    # pprint(raw_table)

    keys = raw_table.keys()
    rows = list(range(2, ws.max_row+1))

    for i in range(len(keys)):
        row = [1] + list(ws[rows[i]])
        row[1:] = raw_table[schoolDays[i]]
        p = ws[rows[i]]
        for j in range(1, len(p)):
            p[j].value = row[j]

    style_worksheet(ws, f'A1:{colLetters[ws.max_column - 1]}{ws.max_row}')

    wb.save(rf'{folder}/{name}.xlsx')


teachersAndSubjects = {
    'Maths(SU)': 'Susanna Abraham',
    'Maths(GP)': 'Ganesaperumal B',
    'Chemistry(B)': 'Bini P Kuriakose',
    'English(B)': 'Bhaswati Chattopadhyay',
    'Computer(J)': 'Jones Solomon Roche',
    'Biology(S)': 'Swami',
    'Physics(SS)': 'Susan Sobi',
    'PT(M)': 'Maruthupandian'
} #Test data

# createNewTimetable(
#     8, 5, r"C:\Users\abhin\OneDrive\Desktop\do NOT delete", 'random', teachersAndSubjects)
